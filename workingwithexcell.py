import openpyxl
from openpyxl import load_workbook

wbinput = load_workbook('Input.xlsx')
wboutput = load_workbook('Output.xlsx')

#print(wbinput.sheetnames)

sheet_input = wbinput['Sheet1']
Output_Sheetname = input("Name of the sheet to be edited : ")
sheet_output = wboutput[Output_Sheetname]

column_no = int(input("Unique Key Column Number : "))

for k in range(2,sheet_input.max_row):
    for l in range(2, sheet_output.max_row):
            if(sheet_input.cell(k,column_no).value == sheet_output.cell(l,column_no).value):
                for i in range (1, sheet_input.max_column): 
                       c = sheet_input.cell(row = k, column = i) 
                       sheet_output.cell(row = l, column = i).value = c.value
            continue

print(sheet_input.max_row," Number of rows updated")
wboutput.save('Output2.xlsx')